from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config import EXCEL_FILE, SHEET_MAP
import re

# Human-readable labels for every field name in every form
FIELD_LABELS = {
    # Admission
    "hrn_no": "HRN No.",
    "med_record_no": "Med. Record No.",
    "patient_last": "Patient's Name (Last)",
    "patient_first": "Patient's Name (First)",
    "patient_suffix": "Suffix",
    "patient_middle": "Middle Name",
    "service": "Service",
    "rooms": "Rooms",
    "address": "Permanent Address",
    "tel_no": "Telephone No.",
    "sex": "Sex",
    "civil_status": "Civil Status",
    "birth_date": "Birth Date",
    "age": "Age",
    "birth_place": "Birth Place",
    "nationality": "Nationality",
    "religion": "Religion",
    "occupation": "Occupation",
    "employer": "Employer",
    "employer_address": "Employer Address",
    "employer_tel": "Employer Tel. No.",
    "father": "Father",
    "mother": "Mother",
    "spouse": "Spouse",
    "spouse_address": "Spouse Address",
    "spouse_tel": "Spouse Tel. No.",
    "admission_datetime": "Admission Date and Time",
    "discharge_datetime": "Discharge Date and Time",
    "day_of_stay": "Day of Stay",
    "attending_physician": "Attending Physician",
    "type_of_admission": "Type of Admission",
    "admitted_by": "Admitted by",
    "resident": "Resident",
    "allergic_to": "Allergic to",
    "social_services": "Social Services",
    "hosp_plan": "Hospitalization Plan",
    "health_insurance": "Health Insurance Name",
    "philhealth_id": "PhilHealth ID",
    "data_furnished_by": "Data Furnished by",
    "data_address": "Address",
    "relationship": "Relationship to Patient",
    "admission_dx": "Admission Diagnosis",
    "admission_icd": "Admission ICD CODE No.",
    "principal_dx": "Principal Diagnosis",
    "principal_icd": "Principal ICD CODE No.",
    "no_smoker": "NO SMOKER",
    "smoker": "SMOKER",
    "other_dx": "Other Diagnosis",
    "principal_procedures": "Principal Operating Procedures",
    "other_procedures": "Other Operation/Procedure",
    "accident_injuries": "Accident/Injuries/Poisoning",
    "place_of_occurrence": "Place of Occurrence",
    "disp_discharge": "Disposition: Discharge",
    "disp_transferred": "Disposition: Transferred",
    "disp_absconded": "Disposition: Absconded",
    "disp_dama": "Disposition: DAMA",
    "result_recovered": "Result: Recovered",
    "result_improved": "Result: Improved",
    "result_died": "Result: Died",
    "result_unimproved": "Result: Unimproved",
    "result_48h_minus": "Result: -48 Hours",
    "result_48h_plus": "Result: +48 Hours",
    "result_autopsied": "Result: Autopsied",
    "result_not_autopsied": "Result: Not Autopsied",
    "attending_sig": "Attending Physician Signature",
    "attending_md": "Attending Physician M.D.",
    "audit": "Audit",
    # Vitals
    "patient_name": "Patient Name",
    "ward": "Ward",
    # History
    "history_date": "Date",
    "vs_bp": "Vital Signs - BP",
    "vs_hr": "Vital Signs - HR",
    "vs_rr": "Vital Signs - RR",
    "vs_temp": "Vital Signs - Temperature",
    "gen_awake": "General Survey: Awake and Alert",
    "gen_altered": "General Survey: Altered Sensorium",
    "gen_others": "General Survey: Others",
    "clinical_impression": "Clinical Impression",
    # MIO
    "room_no": "Room No.",
    "mio_date": "Date",
    "admission_weight": "Admission Weight",
    "current_weight": "Current Weight",
    # Medication
    "ward_room": "Ward/Room",
    "physician": "Physician",
    # IV Fluid / Kardex
    "hospital_no": "Hospital No.",
    "tpr": "TPR",
    "mio_val": "MIO",
    "diagnosis": "Diagnosis (Dx)",
    "date_admitted": "Date Admitted",
    "o2_lpm": "O2 Inhalation (LPM)",
    "vr_setup": "VR Set-up",
    "diet": "Diet",
    "cbg": "CBG",
    "to_follow": "To Follow",
}

def _label(key):
    """Convert a field key to a human-readable label."""
    if key in FIELD_LABELS:
        return FIELD_LABELS[key]
    # Strip trailing _N or _N_N patterns for table rows
    clean = re.sub(r'_\d+(_\d+)*$', '', key)
    if clean in FIELD_LABELS:
        return FIELD_LABELS[clean]
    # Fallback: humanize snake_case
    return clean.replace("_", " ").title()

def _row_label(key):
    """For table row keys like bp_1, s1_po_2, med_a_3 — add context."""
    # Vital signs rows: bp_1, cr_2, etc.
    m = re.match(r'^(bp|cr|rr|temp|spo2|sig|date|time)_(\d+)$', key)
    if m:
        field_map = {"bp":"Blood Pressure","cr":"Cardiac Rate","rr":"Respiratory Rate",
                     "temp":"Temperature","spo2":"O2 Saturation","sig":"Signature",
                     "date":"Date","time":"Time"}
        return f"Row {m.group(2)} – {field_map.get(m.group(1), m.group(1).upper())}"

    # MIO rows: s1_po_3, s2_urine_1, etc.
    m = re.match(r'^s(\d)_(po|ivf|tube|oth_in|urine|emesis|ng|stool|oth_out|tot_\w+|na_sig)_?(\d*)$', key)
    if m:
        field_map = {"po":"PO","ivf":"IV Fluids","tube":"Tube Feeding","oth_in":"Other (Intake)",
                     "urine":"Urine","emesis":"Emesis","ng":"NG","stool":"Stool","oth_out":"Other (Output)",
                     "na_sig":"Nursing Asst. Signature"}
        col = m.group(2)
        if col.startswith("tot_"):
            col = "Total – " + field_map.get(col[4:], col[4:])
        else:
            col = field_map.get(col, col)
        row_suffix = f" Row {m.group(3)}" if m.group(3) else ""
        return f"Block {m.group(1)} – {col}{row_suffix}"

    # Medication: med_1, date_ordered_1, hr_1_1_2, sig_1_2_3
    m = re.match(r'^(hr|sig)_(\d+)_(\d+)_(\d+)$', key)
    if m:
        shifts = {1:"11-7", 2:"7-3", 3:"3-11"}
        return f"Medication {m.group(2)}, Shift {shifts.get(int(m.group(3)),'?')}, Col {m.group(4)} – {'Hr' if m.group(1)=='hr' else 'Sig'}"

    m = re.match(r'^med_date_(\d+)_(\d+)$', key)
    if m:
        shifts = {1:"11-7", 2:"7-3", 3:"3-11"}
        return f"Medication {m.group(1)} – Date (Shift {shifts.get(int(m.group(2)),'?')})"

    m = re.match(r'^(date_ordered|med)_(\d+)$', key)
    if m:
        lbl = "Date Ordered" if m.group(1)=="date_ordered" else "Medication"
        return f"Medication {m.group(2)} – {lbl}"

    m = re.match(r'^(nurse_name|nurse_initial)_(\d+)$', key)
    if m:
        return f"Nurse {m.group(2)} – {'Name' if 'name' in m.group(1) else 'Initial'}"

    # IV Fluid rows
    m = re.match(r'^(ivf_type|drug_add|iv_rate|started|consumed|ivf_sig)_(\d+)$', key)
    if m:
        lbl_map = {"ivf_type":"Type of Fluid","drug_add":"Drug Additives","iv_rate":"IV Rate",
                   "started":"Time/Date Started","consumed":"Time/Date Consumed","ivf_sig":"Nurse Signature"}
        return f"Bottle {m.group(2)} – {lbl_map.get(m.group(1), m.group(1))}"

    # Kardex medications
    m = re.match(r'^med_(date_)?(a|b)_(\d+)$', key)
    if m:
        col = "Column A" if m.group(2)=="a" else "Column B"
        what = "Date" if m.group(1) else "Medication"
        return f"Kardex {col} Row {m.group(3)} – {what}"

    m = re.match(r'^(inh_date|inhalation)_(\d+)$', key)
    if m:
        return f"Inhalation {m.group(2)} – {'Date' if 'date' in m.group(1) else 'Medication'}"

    m = re.match(r'^(prn_date|prn_med)_(\d+)$', key)
    if m:
        return f"PRN {m.group(2)} – {'Date' if 'date' in m.group(1) else 'Medication'}"

    m = re.match(r'^(lab_date|lab|lab_rem)_(\d+)$', key)
    if m:
        lbl_map = {"lab_date":"Date","lab":"Lab/Diagnostics","lab_rem":"Remarks"}
        return f"Lab Row {m.group(2)} – {lbl_map.get(m.group(1), m.group(1))}"

    m = re.match(r'^(endorse_date|endorse)_(\d+)$', key)
    if m:
        return f"Endorsement {m.group(2)} – {'Date' if 'date' in m.group(1) else 'Endorsement'}"

    m = re.match(r'^(order_date|order_time|soap|doctors_order)_(\d+)$', key)
    if m:
        lbl_map = {"order_date":"Date","order_time":"Time","soap":"SOAP Notes","doctors_order":"Doctor's Order"}
        return f"Entry {m.group(2)} – {lbl_map.get(m.group(1), m.group(1))}"

    m = re.match(r'^(note_dt|diet|focus|dar)_(\d+)$', key)
    if m:
        lbl_map = {"note_dt":"Date/Time","diet":"Diet","focus":"Focus","dar":"D-A-R"}
        return f"Row {m.group(2)} – {lbl_map.get(m.group(1), m.group(1))}"

    # General fallback
    return _label(key)


def append_submission_to_excel(submission):
    """Write a form submission into the correct Excel sheet as a formatted block."""
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet_name = SHEET_MAP.get(submission["form_id"])
        if not sheet_name:
            return
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]
        start_row = ws.max_row + 2  # blank separator row

        # ── Header banner ─────────────────────────────────
        teal_fill = PatternFill("solid", fgColor="0D7377")
        hdr = ws.cell(row=start_row, column=1,
                      value=f"SUBMISSION — {submission['form_name'].upper()}")
        hdr.font = Font(bold=True, color="FFFFFF", size=11)
        hdr.fill = teal_fill
        hdr.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)

        meta_row = start_row + 1
        meta_items = [
            ("Submitted by", submission["nurse_name"]),
            ("Email", submission["nurse_email"]),
            ("Date/Time", submission["submitted_at"]),
            ("Patient", submission.get("patient_name", "—")),
        ]
        navy_fill = PatternFill("solid", fgColor="0A1628")
        for col_idx, (label, value) in enumerate(meta_items, 1):
            lc = ws.cell(row=meta_row, column=col_idx*2-1, value=label+":")
            lc.font = Font(bold=True, color="FFFFFF", size=9)
            lc.fill = navy_fill
            vc = ws.cell(row=meta_row, column=col_idx*2, value=value)
            vc.font = Font(color="FFFFFF", size=9)
            vc.fill = navy_fill

        # ── Column headers ────────────────────────────────
        hdr_row = meta_row + 1
        col_hdr_fill = PatternFill("solid", fgColor="E8F4F4")
        for col_idx, txt in enumerate(["Field", "Value"], 1):
            c = ws.cell(row=hdr_row, column=col_idx, value=txt)
            c.font = Font(bold=True, color="0A1628", size=9)
            c.fill = col_hdr_fill

        # ── Data rows ─────────────────────────────────────
        light_fill = PatternFill("solid", fgColor="F9FDFD")
        data_row = hdr_row + 1
        zebra = False
        for key, value in submission["data"].items():
            if not value or value in ("No", "false", ""):
                continue
            label = _row_label(key)
            lc = ws.cell(row=data_row, column=1, value=label)
            lc.font = Font(bold=True, size=9, color="0A1628")
            vc = ws.cell(row=data_row, column=2, value=str(value))
            vc.font = Font(size=9)
            vc.alignment = Alignment(wrap_text=True, vertical="top")
            if zebra:
                lc.fill = light_fill
                vc.fill = light_fill
            zebra = not zebra
            data_row += 1

        # ── Column widths ─────────────────────────────────
        ws.column_dimensions[get_column_letter(1)].width = 38
        ws.column_dimensions[get_column_letter(2)].width = 55
        for i in range(3, 9):
            ws.column_dimensions[get_column_letter(i)].width = 22

        wb.save(EXCEL_FILE)

    except Exception as exc:
        import traceback
        traceback.print_exc()
        print(f"[Excel] Write error: {exc}")


def get_excel_file_path():
    return EXCEL_FILE
