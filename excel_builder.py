"""
excel_builder.py — builds a fresh Excel from all submissions (one row per submission).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, json
from config import SUBMISSIONS_FILE

TEAL = "0D7377"; TEAL_LIGHT = "E8F6F6"; NAVY = "0A1628"
WHITE = "FFFFFF"; CREAM = "FAF8F3"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color=NAVY, size=10): return Font(bold=bold, color=color, size=size, name="Calibri")
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _align(wrap=False, center=False):
    return Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=wrap)

def _header_row(ws, col_defs, row=1):
    for ci, (label, width) in enumerate(col_defs, 1):
        c = ws.cell(row=row, column=ci, value=label)
        c.font = _font(bold=True, color=WHITE, size=10)
        c.fill = _fill(TEAL); c.alignment = _align(wrap=True, center=True); c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

def _title_row(ws, title, num_cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = _font(bold=True, color=WHITE, size=12)
    c.fill = _fill(NAVY); c.alignment = _align(center=True)
    ws.row_dimensions[row].height = 22

def _data_row(ws, values, row_num):
    fill = _fill(TEAL_LIGHT) if row_num % 2 == 0 else _fill(CREAM)
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=ci, value=val or "")
        c.font = _font(size=9); c.fill = fill
        c.alignment = _align(wrap=True); c.border = _border()

def _checked(d, keys, other_key=None):
    labels = {
        "gen_awake":"Awake & Alert","gen_altered":"Altered Sensorium",
        "heent_normal":"Essentially Normal","heent_icteric":"Icteric Sclera","heent_dry":"Dry Mucus Membrane",
        "heent_pupil":"Abnormal Pupillary Rxn","heent_pale_conj":"Pale Conjunctive",
        "heent_sunken_font":"Sunken Fontanelle","heent_lymph":"Cervical Lymphadenopathy","heent_sunken_eye":"Sunken Eyeballs",
        "chest_normal":"Essentially Normal","chest_lump":"Lump/s Over Breast","chest_wheeze":"Wheeze",
        "chest_asym":"Asymmetrical Expansion","chest_rales":"Rales/Crackles/Rhonchi",
        "chest_retract":"Intercostal Retractions","chest_decreased":"Decreased Breath Sounds",
        "cvs_normal":"Essentially Normal","cvs_irregular":"Irregular Rhythm","cvs_murmur":"Murmur",
        "cvs_displaced":"Displaced Apex Beat","cvs_muffled":"Muffled Heart Sound",
        "cvs_pericardial":"Pericardial Bulge","cvs_heaves":"Heaves/Thrills",
        "abd_normal":"Essentially Normal","abd_palpable":"Palpable Mass","abd_uterine":"Uterine Contraction",
        "abd_rigidity":"Abdomen Rigidity","abd_tympanic":"Tympanic/Dull Abdomen",
        "abd_tenderness":"Abdominal Tenderness","abd_hyperactive":"Hyperactive Bowel Sounds",
        "gu_normal":"Essentially Normal","gu_blood":"Blood Stained Exam Finger","gu_discharge":"Abnormal Vaginal Discharge",
        "skin_normal":"Essentially Normal","skin_edema":"Edema/Swelling","skin_rashes":"Rashes",
        "skin_poor_turgor":"Poor Skin Turgor","skin_clubbing":"Clubbing","skin_decreased":"Decreased Mobility",
        "skin_weak":"Weak Pulses","skin_cold":"Cold Clammy Skin","skin_pale":"Pale Nailbeds","skin_cyanosis":"Cyanosis/Mottled Skin",
        "neuro_normal":"Essentially Normal","neuro_reflex":"Abnormal Reflex(es)","neuro_gait":"Abnormal Gait",
        "neuro_sensation":"Abnormal/Decreased Sensation","neuro_coord":"Poor Coordination",
        "o2_cannula":"Cannula","o2_face_mask":"Face Mask","o2_nasal_cath":"Nasal Cath",
        "cont_o2":"O2","cont_ett":"ETT","cont_ngt":"NGT/OGT","cont_foley":"Foley Catheter","cont_ctt":"CTT",
    }
    parts = [labels.get(k, k.replace("_"," ").title()) for k in keys if d.get(k)]
    if other_key and d.get(other_key): parts.append(d[other_key])
    return ", ".join(parts) if parts else ""

def _get_sheet_name(form_id):
    from config import SHEET_MAP
    return SHEET_MAP.get(form_id)

SHEETS = {
"Admission and Discharge Record": {
    "title": "eTelemo — ADMISSION AND DISCHARGE RECORD",
    "cols": [("Submitted By",18),("Date Submitted",16),("Nurse Email",22),
        ("HRN No.",12),("Med. Record No.",16),("Last Name",16),("First Name",16),("Suffix",8),("Middle Name",16),
        ("Service",14),("Rooms",10),("Permanent Address",28),("Telephone No.",16),
        ("Sex",8),("Civil Status",14),("Birth Date",14),("Age",6),("Birth Place",16),
        ("Nationality",14),("Religion",14),("Occupation",16),
        ("Employer",20),("Employer Address",24),("Employer Tel.",14),
        ("Father",18),("Mother",18),("Spouse",18),("Spouse Address",24),("Spouse Tel.",14),
        ("Admission Date & Time",20),("Discharge Date & Time",20),("Day of Stay",12),("Attending Physician",24),
        ("Type of Admission",18),("Admitted by",18),("Resident",16),("Allergic to",18),
        ("Social Services",18),("Hospitalization Plan",20),("Health Insurance",20),("PhilHealth ID",16),
        ("Data Furnished by",20),("Data Address",24),("Relationship",16),
        ("Admission Diagnosis",32),("Admission ICD",14),("Principal Diagnosis",32),("Principal ICD",14),
        ("Smoker Status",14),("Other Diagnosis",24),("Principal Procedures",28),
        ("Other Procedures",24),("Accident/Injuries",28),("Place of Occurrence",24),
        ("Disposition",18),("Result",24),("Attending Sig.",20),("Audit",14),],
    "field_map": lambda d: [
        d.get("patient_last",""),d.get("patient_first",""),d.get("patient_suffix",""),d.get("patient_middle",""),
        d.get("service",""),d.get("rooms",""),d.get("address",""),d.get("tel_no",""),
        d.get("sex",""),d.get("civil_status",""),d.get("birth_date",""),d.get("age",""),
        d.get("birth_place",""),d.get("nationality",""),d.get("religion",""),d.get("occupation",""),
        d.get("employer",""),d.get("employer_address",""),d.get("employer_tel",""),
        d.get("father",""),d.get("mother",""),d.get("spouse",""),d.get("spouse_address",""),d.get("spouse_tel",""),
        d.get("admission_datetime",""),d.get("discharge_datetime",""),d.get("day_of_stay",""),d.get("attending_physician",""),
        d.get("type_of_admission",""),d.get("admitted_by",""),d.get("resident",""),d.get("allergic_to",""),
        d.get("social_services",""),d.get("hosp_plan",""),d.get("health_insurance",""),d.get("philhealth_id",""),
        d.get("data_furnished_by",""),d.get("data_address",""),d.get("relationship",""),
        d.get("admission_dx",""),d.get("admission_icd",""),d.get("principal_dx",""),d.get("principal_icd",""),
        ("NO SMOKER" if d.get("no_smoker") else "")+("SMOKER" if d.get("smoker") else ""),
        d.get("other_dx",""),d.get("principal_procedures",""),d.get("other_procedures",""),
        d.get("accident_injuries",""),d.get("place_of_occurrence",""),
        ", ".join(k.replace("disp_","").title() for k in d if k.startswith("disp_") and d[k]),
        ", ".join(k.replace("result_","").replace("_"," ").title() for k in d if k.startswith("result_") and d[k]),
        d.get("attending_sig",""),d.get("audit",""),],
},
"vs": {
    "title": "eTELEmo — VITAL SIGNS MONITORING SHEET",
    "cols": [("Submitted By",18),("Date Submitted",16),("Nurse Email",22),
        ("Patient Name",22),("Age",6),("Sex",8),("Ward",12),("Row#",6),
        ("Date",14),("Time",10),("Blood Pressure",16),("Cardiac Rate",14),
        ("Respiratory Rate",16),("Temperature",14),("O2 Saturation (%)",16),("Signature",20),],
    "multi_row": True,
    "row_prefix": lambda d: [d.get("patient_name",""),d.get("age",""),d.get("sex",""),d.get("ward","")],
    "row_keys": ["date","time","bp","cr","rr","temp","spo2","sig"],
    "num_rows": 50,
},
"HISTORY RECORD": {
    "title": "eTELEmo — HISTORY RECORD",
    "cols": [("Submitted By",18),("Date Submitted",16),("Nurse Email",22),
        ("Patient Name",22),("Date",14),("BP",10),("HR",10),("RR",10),("Temperature",14),
        ("General Survey",28),("HEENT",40),("Chest/Lungs",40),("CVs",36),("Abdomen",40),
        ("Fundic Height",14),("FHT",12),("GU (IE)",36),("Skin/Extremities",40),("Neuro-Exam",36),
        ("Others",36),("Clinical Impression",40),("Attending Physician",24),],
    "field_map": lambda d: [
        d.get("patient_name",""),d.get("history_date",""),
        d.get("vs_bp",""),d.get("vs_hr",""),d.get("vs_rr",""),d.get("vs_temp",""),
        _checked(d,["gen_awake","gen_altered"],"gen_others"),
        _checked(d,["heent_normal","heent_icteric","heent_dry","heent_pupil","heent_pale_conj","heent_sunken_font","heent_lymph","heent_sunken_eye"],"heent_others"),
        _checked(d,["chest_normal","chest_lump","chest_wheeze","chest_asym","chest_rales","chest_retract","chest_decreased"],"chest_others"),
        _checked(d,["cvs_normal","cvs_irregular","cvs_murmur","cvs_displaced","cvs_muffled","cvs_pericardial","cvs_heaves"],"cvs_others"),
        _checked(d,["abd_normal","abd_palpable","abd_uterine","abd_rigidity","abd_tympanic","abd_tenderness","abd_hyperactive"],"abd_others"),
        d.get("abd_fundic",""),d.get("abd_fht",""),
        _checked(d,["gu_normal","gu_blood","gu_discharge"])+(" | IE: "+d["gu_ie"] if d.get("gu_ie") else ""),
        _checked(d,["skin_normal","skin_edema","skin_rashes","skin_poor_turgor","skin_clubbing","skin_decreased","skin_weak","skin_cold","skin_pale","skin_cyanosis"],"skin_others"),
        _checked(d,["neuro_normal","neuro_reflex","neuro_gait","neuro_sensation","neuro_coord"],"neuro_others"),
        d.get("other_findings",""),d.get("clinical_impression",""),d.get("attending_physician",""),],
},
"MIO": {
    "title": "eTELEmo — INTAKE & OUTPUT (MIO)",
    "cols": [("Submitted By",18),("Date Submitted",16),("Nurse Email",22),
        ("Patient Name",22),("Room #",10),("Date",14),("Admission Weight",16),("Current Weight",16),
        ("Block",8),("Row#",6),("Time",10),
        ("PO",10),("IV Fluids",12),("Tube Feeding",14),("Other (In)",12),
        ("Urine",10),("Emesis",10),("NG",8),("Stool",10),("Other (Out)",12),],
    "multi_row": True, "blocks": True,
    "row_prefix": lambda d: [d.get("patient_name",""),d.get("room_no",""),d.get("mio_date",""),d.get("admission_weight",""),d.get("current_weight","")],
    "row_keys_mio": ["po","ivf","tube","oth_in","urine","emesis","ng","stool","oth_out"],
    "num_rows": 20,
},
"Medication sheet": {
    "title": "eTELEmo — MEDICATION SHEET",
    "cols": [("Submitted By",18),("Date Submitted",16),("Nurse Email",22),
        ("Patient Name",22),("Ward/Room",14),("Physician",20),
        ("Med #",6),("Date Ordered",14),("Medication / Dosage / Frequency",36),
        ("Shift",8),("Date",12),
        ("Hr1",6),("Sig1",10),("Hr2",6),("Sig2",10),("Hr3",6),("Sig3",10),
        ("Hr4",6),("Sig4",10),("Hr5",6),("Sig5",10),("Hr6",6),("Sig6",10),],
    "multi_row": True, "med_rows": True,
    "row_prefix": lambda d: [d.get("patient_name",""),d.get("ward_room",""),d.get("physician","")],
    "num_meds": 20,
},
"Iv fluid chart": {
    "title": "eTELEmo — IV FLUID CHART",
    "cols": [("Submitted By",18),("Date Submitted",16),("Nurse Email",22),
        ("HRN",12),("Patient Name",22),("Ward/Room",14),("Physician",20),
        ("Bottle #",8),("Type of Fluid",20),("Drug Additives",20),("IV Rate",10),
        ("Time/Date Started",20),("Time/Date Consumed",20),("Nurses Signature Over Printed Name",32),],
    "multi_row": True, "ivf_rows": True,
    "row_prefix": lambda d: [d.get("hrn_no",""),d.get("patient_name",""),d.get("ward_room",""),d.get("physician","")],
    "num_bottles": 20,
},
"Kardex": {
    "title": "eTELEmo — KARDEX",
    "cols": [("Submitted By",18),("Date Submitted",16),("Nurse Email",22),
        ("Patient Name",22),("Hospital No.",16),("Age",6),("Sex",8),("Religion",14),("Civil Status",14),
        ("Rm/Ward",12),("TPR",12),("MIO",12),("Dx",32),("Attending Physician",24),("Date Admitted",14),
        ("VR Set-up",20),("Diet",16),("CBG",10),("O2 LPM",8),("O2 Device",20),("Contraption List",36),
        ("Medications Col A",40),("Medications Col B",40),("Inhalations",36),("PRN Medications",36),
        ("Present/Started/Due",36),("IV Fluids",36),("Side Drips",36),
        ("Lab/Diagnostics",40),("Special Endorsements",40),],
    "field_map": lambda d: [
        d.get("patient_name",""),d.get("hospital_no",""),d.get("age",""),
        ("M" if d.get("sex_m") else "")+("F" if d.get("sex_f") else ""),
        d.get("religion",""),d.get("civil_status",""),d.get("room_ward",""),d.get("tpr",""),d.get("mio_val",""),
        d.get("diagnosis",""),d.get("physician",""),d.get("date_admitted",""),
        d.get("vr_setup",""),d.get("diet",""),d.get("cbg",""),d.get("o2_lpm",""),
        _checked(d,["o2_cannula","o2_face_mask","o2_nasal_cath"]),
        _checked(d,["cont_o2","cont_ett","cont_ngt","cont_foley","cont_ctt"])+(" "+d["cont_others"] if d.get("cont_others") else ""),
        " | ".join(f"{d.get('med_date_a_'+str(i),'')}: {d.get('med_a_'+str(i),'')}" for i in range(1,9) if d.get("med_a_"+str(i))),
        " | ".join(f"{d.get('med_date_b_'+str(i),'')}: {d.get('med_b_'+str(i),'')}" for i in range(1,5) if d.get("med_b_"+str(i))),
        " | ".join(f"{d.get('inh_date_'+str(i),'')}: {d.get('inhalation_'+str(i),'')}" for i in range(1,4) if d.get("inhalation_"+str(i))),
        " | ".join(f"{d.get('prn_date_'+str(i),'')}: {d.get('prn_med_'+str(i),'')}" for i in range(1,6) if d.get("prn_med_"+str(i))),
        " | ".join(d.get("present_"+str(i),"") for i in range(1,6) if d.get("present_"+str(i))),
        " | ".join(d.get("ivf_k_"+str(i),"") for i in range(1,4) if d.get("ivf_k_"+str(i))),
        " | ".join(d.get("side_drip_"+str(i),"") for i in range(1,7) if d.get("side_drip_"+str(i))),
        " | ".join(f"{d.get('lab_date_'+str(i),'')}: {d.get('lab_'+str(i),'')} ({d.get('lab_rem_'+str(i),'')})" for i in range(1,9) if d.get("lab_"+str(i))),
        " | ".join(f"{d.get('endorse_date_'+str(i),'')}: {d.get('endorse_'+str(i),'')}" for i in range(1,9) if d.get("endorse_"+str(i))),],
},
"Doctors Order Progress Notes": {
    "title": "eTELEmo — DOCTOR'S ORDER / PROGRESS NOTES",
    "cols": [("Submitted By",18),("Date Submitted",16),("Nurse Email",22),
        ("HRN",12),("Patient Name",22),("Ward/Room",14),("Physician",20),
        ("Entry #",8),("Date",14),("Time",10),("PROGRESS NOTES (S-O-A-P)",50),("DOCTOR'S ORDER",50),],
    "multi_row": True, "order_rows": True,
    "row_prefix": lambda d: [d.get("hrn_no",""),d.get("patient_name",""),d.get("ward_room",""),d.get("physician","")],
    "num_entries": 30,
},
"Nurses Progress Notes": {
    "title": "eTELEmo — NURSES PROGRESS NOTES",
    "cols": [("Submitted By",18),("Date Submitted",16),("Nurse Email",22),
        ("HRN",12),("Patient Name",22),("Ward/Room",14),("Physician",20),
        ("Row#",6),("Date/Time",18),("Diet",14),("Focus",16),("D=Data/A=Action/R=Response",60),],
    "multi_row": True, "notes_rows": True,
    "row_prefix": lambda d: [d.get("hrn_no",""),d.get("patient_name",""),d.get("ward_room",""),d.get("physician","")],
    "num_rows": 40,
},
"vs graph sheet": {
    "title": "eTELEmo — VS GRAPH SHEET",
    "cols": [("Submitted By",18),("Date Submitted",16),("Nurse Email",22),
        ("Patient Name",22),("Graph Date",14),("Days of Hospitalization",22),("Post-Operative Days",20),
        ("Section",12),("Day",8),("AM/PM",8),("Time Slot",10),("Scale Value",12),("Reading",12),],
    "multi_row": True, "vs_graph": True,
    "row_prefix": lambda d: [d.get("patient_name",""),d.get("graph_date",""),d.get("days_hosp",""),d.get("post_op_days","")],
},
}


def build_excel(output_path):
    wb = Workbook()
    wb.remove(wb.active)

    subs = []
    if os.path.exists(SUBMISSIONS_FILE):
        with open(SUBMISSIONS_FILE) as f:
            subs = json.load(f)

    for sheet_name, sheet_def in SHEETS.items():
        sheet_subs = [s for s in subs if _get_sheet_name(s["form_id"]) == sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        cols = sheet_def["cols"]
        num_cols = len(cols)

        _title_row(ws, sheet_def["title"], num_cols, row=1)
        ws.row_dimensions[1].height = 22
        _header_row(ws, cols, row=2)
        ws.row_dimensions[2].height = 32
        ws.freeze_panes = "A3"

        data_row = 3

        for sub in sheet_subs:
            d = sub.get("data", {})
            meta = [sub["nurse_name"], sub["submitted_at"], sub["nurse_email"]]

            if sheet_def.get("multi_row"):

                if sheet_def.get("row_keys"):
                    # Vitals
                    prefix = sheet_def["row_prefix"](d)
                    for i in range(1, sheet_def["num_rows"]+1):
                        row_vals = [d.get(f"{k}_{i}","") for k in sheet_def["row_keys"]]
                        if any(v for v in row_vals):
                            _data_row(ws, meta + prefix + [i] + row_vals, data_row)
                            data_row += 1

                elif sheet_def.get("med_rows"):
                    # Medication
                    prefix = sheet_def["row_prefix"](d)
                    for i in range(1, sheet_def["num_meds"]+1):
                        med = d.get(f"med_{i}","")
                        date_ord = d.get(f"date_ordered_{i}","")
                        if not med: continue
                        for si, shift in enumerate(["11-7","7-3","3-11"], 1):
                            hr_sigs = []
                            for j in range(1,7):
                                hr_sigs += [d.get(f"hr_{i}_{si}_{j}",""), d.get(f"sig_{i}_{si}_{j}","")]
                            row_vals = [i, date_ord, med, shift, d.get(f"med_date_{i}_{si}","")] + hr_sigs
                            _data_row(ws, meta + prefix + row_vals, data_row)
                            data_row += 1

                elif sheet_def.get("ivf_rows"):
                    # IV Fluid
                    prefix = sheet_def["row_prefix"](d)
                    for i in range(1, sheet_def["num_bottles"]+1):
                        if not d.get(f"ivf_type_{i}"): continue
                        row_vals = [i, d.get(f"ivf_type_{i}",""), d.get(f"drug_add_{i}",""),
                                    d.get(f"iv_rate_{i}",""), d.get(f"started_{i}",""),
                                    d.get(f"consumed_{i}",""), d.get(f"ivf_sig_{i}","")]
                        _data_row(ws, meta + prefix + row_vals, data_row)
                        data_row += 1

                elif sheet_def.get("order_rows"):
                    # Doctor's Order
                    prefix = sheet_def["row_prefix"](d)
                    for i in range(1, sheet_def["num_entries"]+1):
                        soap = d.get(f"soap_{i}","")
                        order = d.get(f"doctors_order_{i}","")
                        if not soap and not order: continue
                        row_vals = [i, d.get(f"order_date_{i}",""), d.get(f"order_time_{i}",""), soap, order]
                        _data_row(ws, meta + prefix + row_vals, data_row)
                        data_row += 1

                elif sheet_def.get("notes_rows"):
                    # Nurses Notes
                    prefix = sheet_def["row_prefix"](d)
                    for i in range(1, sheet_def["num_rows"]+1):
                        dar = d.get(f"dar_{i}","")
                        dt  = d.get(f"note_dt_{i}","")
                        if not dar and not dt: continue
                        row_vals = [i, dt, d.get(f"diet_{i}",""), d.get(f"focus_{i}",""), dar]
                        _data_row(ws, meta + prefix + row_vals, data_row)
                        data_row += 1

                elif sheet_def.get("blocks"):
                    # MIO
                    prefix = sheet_def["row_prefix"](d)
                    for block in range(1, 4):
                        for r in range(1, sheet_def["num_rows"]+1):
                            keys = sheet_def["row_keys_mio"]
                            row_vals = [d.get(f"s{block}_{k}_{r}","") for k in keys]
                            t = d.get(f"s{block}_time_{r}","")
                            if t or any(v for v in row_vals):
                                _data_row(ws, meta + prefix + [block, r, t] + row_vals, data_row)
                                data_row += 1

                elif sheet_def.get("vs_graph"):
                    # VS Graph Sheet
                    prefix = sheet_def["row_prefix"](d)
                    for section in ["cardiac","pulse","temp"]:
                        for day in range(1, 11):
                            for ap in ["am","pm"]:
                                for t in ["12","8","4"]:
                                    for y in [90,80,85,60,50,40,30,20,10,120,110,100,70]:
                                        val = d.get(f"{section}_{ap}_{t}_d{day}_y{y}","")
                                        if val:
                                            row_vals = [section.upper(), day, ap.upper(), t, str(y), val]
                                            _data_row(ws, meta + prefix + row_vals, data_row)
                                            data_row += 1

            else:
                # Single row forms
                field_vals = sheet_def["field_map"](d)
                _data_row(ws, meta + field_vals, data_row)
                data_row += 1

        if data_row == 3:
            ws.cell(row=3, column=1, value="(No submissions yet)").font = _font(color="999999", size=9)

    wb.save(output_path)
    return output_path
